"""Build an auditable Excel DCF model for Netflix that reconciles with the
Python valuation engine.

Sheets:
  Summary  - key outputs, valuation range, football-field chart
  DCF      - three scenario blocks (Bear/Base/Bull) + two sensitivity tables
  WACC     - CAPM cost of capital
  Comps    - peer trading multiples and implied value

Conventions (industry standard):
  blue text  = hardcoded input you can change
  black text = formula / calculation
  green text = link to another sheet
  All calculations are Excel formulas (nothing hardcoded that should compute),
  so the workbook recalculates if you change an input.

Usage:
    python build_excel_model.py
Then validate with the xlsx skill's recalc script.
"""
from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

ROOT = Path(__file__).resolve().parent
sys.path.insert(0, str(ROOT))

import yaml  # noqa: E402
from data.build_db import build, DB_PATH  # noqa: E402
from src.data_loader import get_peer_tickers, get_snapshot  # noqa: E402

OUT = ROOT / "outputs" / "NFLX_DCF_Model.xlsx"

# --- single source of truth: same config + database as the Python engine ---
if not DB_PATH.exists():
    build()
CFG = yaml.safe_load((ROOT / "config" / "netflix.yaml").read_text())
NFLX = get_snapshot("NFLX")
PEER_SNAPS = [get_snapshot(t) for t in get_peer_tickers("NFLX")]
SCN = CFG["scenarios"]

rf = CFG["wacc"]["risk_free_rate"]
erp = CFG["wacc"]["equity_risk_premium"]
kd_pre = CFG["wacc"]["pre_tax_cost_of_debt"]
tax = CFG["tax_rate"]
beta, price, shares = NFLX.beta, NFLX.price, NFLX.shares_outstanding
debt, cash, base_rev = NFLX.total_debt, NFLX.cash, NFLX.revenue
base_tg = CFG["terminal_growth_rate"]
bear_w, bear_tg = SCN["bear"]["wacc_override"], SCN["bear"]["terminal_growth_override"]
bull_w, bull_tg = SCN["bull"]["wacc_override"], SCN["bull"]["terminal_growth_override"]

# ---- styles -------------------------------------------------------------
BLUE = "0000FF"
BLACK = "000000"
GREEN = "008000"
FONT = "Arial"

YELLOW = PatternFill("solid", fgColor="FFF2CC")
HDR_FILL = PatternFill("solid", fgColor="1F2937")   # dark slate
SUB_FILL = PatternFill("solid", fgColor="E5E7EB")    # light grey
thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

COLS = ["D", "E", "F", "G", "H"]          # five projection-year columns
YEARS = ["2026", "2027", "2028", "2029", "2030"]

MONEY = '#,##0;(#,##0);"-"'
PERS = '$#,##0.00'
PRICE = '$#,##0.00'
PCT = '0.0%'
MULT = '0.0x'
FACT = '0.000'


def cell(ws, ref, value=None, *, color=BLACK, bold=False, numfmt=None,
         fill=None, comment=None, align=None, size=10, italic=False, border=False):
    c = ws[ref]
    if value is not None:
        c.value = value
    c.font = Font(name=FONT, size=size, bold=bold, italic=italic, color=color)
    if numfmt:
        c.number_format = numfmt
    if fill:
        c.fill = fill
    if align:
        c.alignment = Alignment(horizontal=align)
    if comment:
        c.comment = Comment(comment, "model")
    if border:
        c.border = BORDER
    return c


def title(ws, ref, text, span_to=None, size=13):
    cell(ws, ref, text, color="FFFFFF", bold=True, size=size, fill=HDR_FILL)
    if span_to:
        col0 = ws[ref].column
        row = ws[ref].row
        for col in range(col0 + 1, ws[span_to].column + 1):
            ws.cell(row=row, column=col).fill = HDR_FILL


def section(ws, ref, text, span_to):
    cell(ws, ref, text, bold=True, fill=SUB_FILL)
    col0 = ws[ref].column
    row = ws[ref].row
    for col in range(col0 + 1, ws[span_to].column + 1):
        ws.cell(row=row, column=col).fill = SUB_FILL


def year_header(ws, row):
    for col, yr in zip(COLS, YEARS):
        cell(ws, f"{col}{row}", yr, bold=True, align="right")


# =========================================================================
wb = Workbook()

# ------------------------------------------------------------------ WACC
ws = wb.active
ws.title = "WACC"
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 30
ws.column_dimensions["B"].width = 4
ws.column_dimensions["C"].width = 14

title(ws, "A1", "WACC — Cost of Capital", span_to="C1")
section(ws, "A3", "Cost of equity (CAPM)", "C3")
cell(ws, "A4", "Risk-free rate (10Y UST)")
cell(ws, "C4", rf, color=BLUE, numfmt=PCT,
     comment="Source: ~10Y US Treasury yield, early 2026.")
cell(ws, "A5", "Beta (5y)")
cell(ws, "C5", beta, color=BLUE, numfmt="0.00",
     comment="Source: NFLX 5y levered beta, ~1.2.")
cell(ws, "A6", "Equity risk premium")
cell(ws, "C6", erp, color=BLUE, numfmt=PCT, comment="Market standard 5.0-6.0%.")
cell(ws, "A7", "Cost of equity")
cell(ws, "C7", "=C4+C5*C6", numfmt=PCT, bold=True)

section(ws, "A9", "Cost of debt", "C9")
cell(ws, "A10", "Pre-tax cost of debt")
cell(ws, "C10", kd_pre, color=BLUE, numfmt=PCT, comment="Approx. yield on NFLX notes.")
cell(ws, "A11", "Tax rate")
cell(ws, "C11", tax, color=BLUE, numfmt=PCT,
     comment="US statutory; NFLX effective rate has run lower.")
cell(ws, "A12", "After-tax cost of debt")
cell(ws, "C12", "=C10*(1-C11)", numfmt=PCT)

section(ws, "A14", "Capital structure", "C14")
cell(ws, "A15", "Share price ($)")
cell(ws, "C15", price, color=BLUE, numfmt=PRICE, fill=YELLOW,
     comment="REFRESH to live quote. Post 10:1 split (Nov-2025) reference.")
cell(ws, "A16", "Shares outstanding (mm)")
cell(ws, "C16", shares, color=BLUE, numfmt=MONEY, comment="Post 10:1 split, ~4.22bn.")
cell(ws, "A17", "Market capitalisation ($mm)")
cell(ws, "C17", "=C15*C16", numfmt=MONEY)
cell(ws, "A18", "Total debt ($mm)")
cell(ws, "C18", debt, color=BLUE, numfmt=MONEY, comment="Source: NFLX FY2025 balance sheet.")
cell(ws, "A19", "Cash & equivalents ($mm)")
cell(ws, "C19", cash, color=BLUE, numfmt=MONEY, comment="Source: NFLX FY2025 balance sheet.")
cell(ws, "A20", "Net debt ($mm)")
cell(ws, "C20", "=C18-C19", numfmt=MONEY)
cell(ws, "A21", "Equity weight")
cell(ws, "C21", "=C17/(C17+MAX(C20,0))", numfmt=PCT)
cell(ws, "A22", "Debt weight")
cell(ws, "C22", "=MAX(C20,0)/(C17+MAX(C20,0))", numfmt=PCT)
cell(ws, "A24", "WACC", bold=True)
cell(ws, "C24", "=C7*C21+C12*C22", numfmt=PCT, bold=True, fill=YELLOW)

# ------------------------------------------------------------------- DCF
d = wb.create_sheet("DCF")
d.sheet_view.showGridLines = False
d.column_dimensions["A"].width = 26
d.column_dimensions["B"].width = 24
for col in "CDEFGH":
    d.column_dimensions[col].width = 13

title(d, "A1", "Netflix, Inc. (NFLX) — DCF Model", span_to="H1")
cell(d, "A2", "FCF-margin method · 5-year explicit forecast · mid-year convention · $mm",
     italic=True, size=9)

section(d, "A4", "Assumptions", "H4")
cell(d, "A5", "Base-year revenue FY2025 ($mm)")
cell(d, "C5", base_rev, color=BLUE, numfmt=MONEY,
     comment="Source: Netflix FY2025 shareholder letter, Jan-2026.")
cell(d, "A6", "Net debt ($mm)")
cell(d, "C6", "=WACC!C20", color=GREEN, numfmt=MONEY)
cell(d, "A7", "Shares outstanding (mm)")
cell(d, "C7", "=WACC!C16", color=GREEN, numfmt=MONEY)
cell(d, "A8", "Reference price ($)")
cell(d, "C8", "=WACC!C15", color=GREEN, numfmt=PRICE)

# revenue growth by case
section(d, "A10", "Revenue growth by case", "H10")
year_header(d, 10)
growth = {11: ("Bear", SCN["bear"]["revenue_growth"]),
          12: ("Base", SCN["base"]["revenue_growth"]),
          13: ("Bull", SCN["bull"]["revenue_growth"])}
for row, (name, vals) in growth.items():
    cell(d, f"A{row}", name)
    for col, v in zip(COLS, vals):
        cell(d, f"{col}{row}", v, color=BLUE, numfmt=PCT)

# fcf margin by case
section(d, "A15", "FCF margin by case (% of revenue)", "H15")
year_header(d, 15)
margin = {16: ("Bear", SCN["bear"]["fcf_margin"]),
          17: ("Base", SCN["base"]["fcf_margin"]),
          18: ("Bull", SCN["bull"]["fcf_margin"])}
for row, (name, vals) in margin.items():
    cell(d, f"A{row}", name)
    for col, v in zip(COLS, vals):
        cell(d, f"{col}{row}", v, color=BLUE, numfmt=PCT)

# wacc & terminal growth by case
section(d, "A20", "Discount rate & terminal growth by case", "H20")
cell(d, "C20", "WACC", bold=True, align="right", fill=SUB_FILL)
cell(d, "D20", "Terminal g", bold=True, align="right", fill=SUB_FILL)
cell(d, "A21", "Bear")
cell(d, "C21", bear_w, color=BLUE, numfmt=PCT, comment="Higher risk premium in bear case.")
cell(d, "D21", bear_tg, color=BLUE, numfmt=PCT)
cell(d, "A22", "Base")
cell(d, "C22", "=WACC!C24", color=GREEN, numfmt=PCT)
cell(d, "D22", base_tg, color=BLUE, numfmt=PCT, comment="GDP-aligned; must be < WACC.")
cell(d, "A23", "Bull")
cell(d, "C23", bull_w, color=BLUE, numfmt=PCT, comment="Lower cost of capital as NFLX de-risks.")
cell(d, "D23", bull_tg, color=BLUE, numfmt=PCT)

# discount period row (mid-year)
cell(d, "A25", "Discount period (mid-year)")
cell(d, "D25", 0.5, color=BLUE, numfmt="0.0", comment="Mid-year convention: t-0.5.")
for i in range(1, 5):
    cell(d, f"{COLS[i]}25", f"={COLS[i-1]}25+1", numfmt="0.0")

# ---- three scenario blocks ----
blocks = [
    {"name": "BEAR CASE", "hdr": 27, "g": 11, "m": 16, "w": "$C$21", "tg": "$D$21"},
    {"name": "BASE CASE", "hdr": 41, "g": 12, "m": 17, "w": "$C$22", "tg": "$D$22"},
    {"name": "BULL CASE", "hdr": 55, "g": 13, "m": 18, "w": "$C$23", "tg": "$D$23"},
]
pershare_cells = {}
for blk in blocks:
    h = blk["hdr"]
    yr, rev, fcf, df_, pv = h + 1, h + 2, h + 3, h + 4, h + 5
    sumpv, tv, pvtv, ev, eq, ps, up = h + 6, h + 7, h + 8, h + 9, h + 10, h + 11, h + 12
    g, m, w, tg = blk["g"], blk["m"], blk["w"], blk["tg"]

    title(d, f"A{h}", blk["name"], span_to=f"H{h}", size=11)
    year_header(d, yr)
    cell(d, f"A{rev}", "Revenue ($mm)")
    cell(d, f"A{fcf}", "Free cash flow ($mm)")
    cell(d, f"A{df_}", "Discount factor")
    cell(d, f"A{pv}", "PV of FCF ($mm)")
    for i, col in enumerate(COLS):
        if i == 0:
            cell(d, f"{col}{rev}", f"=$C$5*(1+{col}${g})", numfmt=MONEY)
        else:
            cell(d, f"{col}{rev}", f"={COLS[i-1]}{rev}*(1+{col}${g})", numfmt=MONEY)
        cell(d, f"{col}{fcf}", f"={col}{rev}*{col}${m}", numfmt=MONEY)
        cell(d, f"{col}{df_}", f"=1/(1+{w})^{col}$25", numfmt=FACT)
        cell(d, f"{col}{pv}", f"={col}{fcf}*{col}{df_}", numfmt=MONEY)

    cell(d, f"B{sumpv}", "Sum PV of explicit FCF")
    cell(d, f"C{sumpv}", f"=SUM(D{pv}:H{pv})", numfmt=MONEY)
    cell(d, f"B{tv}", "Terminal value (Gordon)")
    cell(d, f"C{tv}", f"=H{fcf}*(1+{tg})/({w}-{tg})", numfmt=MONEY)
    cell(d, f"B{pvtv}", "PV of terminal value")
    cell(d, f"C{pvtv}", f"=C{tv}*H{df_}", numfmt=MONEY)
    cell(d, f"B{ev}", "Enterprise value")
    cell(d, f"C{ev}", f"=C{sumpv}+C{pvtv}", numfmt=MONEY, bold=True)
    cell(d, f"B{eq}", "Less: net debt  →  Equity value")
    cell(d, f"C{eq}", f"=C{ev}-$C$6", numfmt=MONEY)
    cell(d, f"B{ps}", "Value per share ($)")
    cell(d, f"C{ps}", f"=C{eq}/$C$7", numfmt=PERS, bold=True, fill=YELLOW)
    cell(d, f"B{up}", "Upside / (downside) vs price")
    cell(d, f"C{up}", f"=C{ps}/$C$8-1", numfmt=PCT)
    pershare_cells[blk["name"]] = f"C{ps}"

# ---- sensitivity 1: WACC (rows) x terminal growth (cols), base FCF path ----
BASE = blocks[1]
base_fcf_row = BASE["hdr"] + 3            # base FCF row -> 44
section(d, "A70", "Sensitivity — base-case value/share: WACC (down) × terminal g (across)", "H70")
# column headers (terminal growth), centred on base
cell(d, "F71", "=$D$22", color=BLACK, numfmt=PCT, bold=True, align="right")
cell(d, "E71", "=F71-0.005", numfmt=PCT, align="right")
cell(d, "D71", "=E71-0.005", numfmt=PCT, align="right")
cell(d, "G71", "=F71+0.005", numfmt=PCT, align="right")
cell(d, "H71", "=G71+0.005", numfmt=PCT, align="right")
# row headers (WACC), centred on base
cell(d, "C74", "=$C$22", numfmt=PCT, bold=True)
cell(d, "C73", "=C74-0.01", numfmt=PCT)
cell(d, "C72", "=C73-0.01", numfmt=PCT)
cell(d, "C75", "=C74+0.01", numfmt=PCT)
cell(d, "C76", "=C75+0.01", numfmt=PCT)
cell(d, "B71", "WACC ↓ / g →", italic=True, size=9)
for r in range(72, 77):
    for col in COLS:
        f = (f"=(SUMPRODUCT($D${base_fcf_row}:$H${base_fcf_row},"
             f"1/(1+$C{r})^($D$25:$H$25))"
             f"+$H${base_fcf_row}*(1+{col}$71)/($C{r}-{col}$71)/(1+$C{r})^$H$25"
             f"-$C$6)/$C$7")
        cell(d, f"{col}{r}", f, numfmt=PERS, border=True)

# ---- sensitivity 2: revenue growth (flat) x FCF margin (flat) ----
section(d, "A79", "Sensitivity — value/share: flat revenue growth (down) × flat FCF margin (across)", "H79")
cell(d, "F80", 0.25, color=BLUE, numfmt=PCT, bold=True, align="right")
cell(d, "E80", "=F80-0.02", numfmt=PCT, align="right")
cell(d, "D80", "=E80-0.02", numfmt=PCT, align="right")
cell(d, "G80", "=F80+0.02", numfmt=PCT, align="right")
cell(d, "H80", "=G80+0.02", numfmt=PCT, align="right")
cell(d, "C83", 0.10, color=BLUE, numfmt=PCT, bold=True)
cell(d, "C82", "=C83-0.02", numfmt=PCT)
cell(d, "C81", "=C82-0.02", numfmt=PCT)
cell(d, "C84", "=C83+0.02", numfmt=PCT)
cell(d, "C85", "=C84+0.02", numfmt=PCT)
cell(d, "B80", "growth ↓ / margin →", italic=True, size=9)
for r in range(81, 86):
    for col in COLS:
        f = (f"=({col}$80*$C$5*SUMPRODUCT((1+$C{r})^($D$25:$H$25+0.5),"
             f"1/(1+$C$22)^($D$25:$H$25))"
             f"+{col}$80*$C$5*(1+$C{r})^5*(1+$D$22)/($C$22-$D$22)/(1+$C$22)^$H$25"
             f"-$C$6)/$C$7")
        cell(d, f"{col}{r}", f, numfmt=PERS, border=True)

cell(d, "A88", "Note: flat-rate table holds WACC & terminal g at base; centre cell of the "
              "top table equals the base case.", italic=True, size=9)

# ------------------------------------------------------------------ Comps
c = wb.create_sheet("Comps")
c.sheet_view.showGridLines = False
c.column_dimensions["A"].width = 12
for col in "BCDEFGHIJKL":
    c.column_dimensions[col].width = 11
for col in "MNOP":
    c.column_dimensions[col].width = 10

title(c, "A1", "Comparable Company Analysis", span_to="P1")
cell(c, "A2", "Peer multiples are ILLUSTRATIVE placeholders — refresh before citing. $mm.",
     italic=True, size=9)
headers = ["Ticker", "Revenue", "EBIT", "D&A", "EBITDA", "Net inc.", "Price",
           "Shares", "Debt", "Cash", "Mkt cap", "EV", "EV/Sales", "EV/EBITDA",
           "EV/EBIT", "P/E"]
for i, htext in enumerate(headers):
    col = chr(ord("A") + i)
    cell(c, f"{col}3", htext, bold=True, align="right" if i else "left",
         fill=SUB_FILL, border=True)

# NFLX row (links where the value already exists elsewhere)
cell(c, "A4", "NFLX", bold=True)
cell(c, "B4", "=DCF!$C$5", color=GREEN, numfmt=MONEY)
cell(c, "C4", NFLX.operating_income, color=BLUE, numfmt=MONEY,
     comment="Source: NFLX FY2025 operating income.")
cell(c, "D4", NFLX.d_and_a, color=BLUE, numfmt=MONEY,
     comment="Estimate: content + PP&E amortisation (dominated by content).")
cell(c, "F4", NFLX.net_income, color=BLUE, numfmt=MONEY,
     comment="Source: NFLX FY2025 net income.")
cell(c, "G4", "=WACC!$C$15", color=GREEN, numfmt=PRICE)
cell(c, "H4", "=WACC!$C$16", color=GREEN, numfmt=MONEY)
cell(c, "I4", "=WACC!$C$18", color=GREEN, numfmt=MONEY)
cell(c, "J4", "=WACC!$C$19", color=GREEN, numfmt=MONEY)

for i, p in enumerate(PEER_SNAPS):
    row = 5 + i
    cell(c, f"A{row}", p.ticker)
    cell(c, f"B{row}", p.revenue, color=BLUE, numfmt=MONEY)
    cell(c, f"C{row}", p.operating_income, color=BLUE, numfmt=MONEY)
    cell(c, f"D{row}", p.d_and_a, color=BLUE, numfmt=MONEY)
    cell(c, f"F{row}", p.net_income, color=BLUE, numfmt=MONEY)
    cell(c, f"G{row}", p.price, color=BLUE, numfmt=PRICE)
    cell(c, f"H{row}", p.shares_outstanding, color=BLUE, numfmt=MONEY)
    cell(c, f"I{row}", p.total_debt, color=BLUE, numfmt=MONEY)
    cell(c, f"J{row}", p.cash, color=BLUE, numfmt=MONEY)

for row in range(4, 10):
    cell(c, f"E{row}", f"=C{row}+D{row}", numfmt=MONEY)            # EBITDA
    cell(c, f"K{row}", f"=G{row}*H{row}", numfmt=MONEY)           # mkt cap
    cell(c, f"L{row}", f"=K{row}+I{row}-J{row}", numfmt=MONEY)    # EV
    cell(c, f"M{row}", f"=L{row}/B{row}", numfmt=MULT)
    cell(c, f"N{row}", f"=L{row}/E{row}", numfmt=MULT)
    cell(c, f"O{row}", f"=L{row}/C{row}", numfmt=MULT)
    cell(c, f"P{row}", f'=IF(F{row}>0,K{row}/F{row},"n/m")', numfmt=MULT)

cell(c, "A11", "Peer median (excl. NFLX)", bold=True)
for col in "MNOP":
    cell(c, f"{col}11", f"=MEDIAN({col}5:{col}9)", numfmt=MULT, bold=True, fill=SUB_FILL)

section(c, "A13", "Implied value — Netflix", "F13")
for i, htext in enumerate(["Method", "Peer median", "NFLX metric", "Implied EV",
                           "Implied equity", "Value/share"]):
    col = chr(ord("A") + i)
    cell(c, f"{col}14", htext, bold=True, align="right" if i else "left", fill=SUB_FILL)
rows_imp = [
    ("EV/Sales", "M11", "B4", 15),
    ("EV/EBITDA", "N11", "E4", 16),
    ("EV/EBIT", "O11", "C4", 17),
]
for name, med, metric, r in rows_imp:
    cell(c, f"A{r}", name)
    cell(c, f"B{r}", f"={med}", numfmt=MULT)
    cell(c, f"C{r}", f"={metric}", numfmt=MONEY)
    cell(c, f"D{r}", f"=B{r}*C{r}", numfmt=MONEY)
    cell(c, f"E{r}", f"=D{r}-WACC!$C$20", numfmt=MONEY)
    cell(c, f"F{r}", f"=E{r}/WACC!$C$16", numfmt=PERS)
cell(c, "A18", "P/E")
cell(c, "B18", "=P11", numfmt=MULT)
cell(c, "C18", "=F4", numfmt=MONEY)
cell(c, "D18", "—", align="right")
cell(c, "F18", "=B18*C18/WACC!$C$16", numfmt=PERS)
cell(c, "A19", "Average implied value/share", bold=True)
cell(c, "F19", "=AVERAGE(F15:F18)", numfmt=PERS, bold=True, fill=YELLOW)

# --------------------------------------------------------------- Summary
s = wb.create_sheet("Summary")
wb.move_sheet("Summary", -(len(wb.sheetnames) - 1))  # move to front
s.sheet_view.showGridLines = False
s.column_dimensions["A"].width = 30
s.column_dimensions["B"].width = 4
s.column_dimensions["C"].width = 14

title(s, "A1", "Netflix, Inc. (NFLX) — Valuation Summary", span_to="C1")
cell(s, "A2", "DCF (FCF-margin) + comparable companies · base year FY2025", italic=True, size=9)

cell(s, "A4", "Reference price ($)")
cell(s, "C4", "=WACC!C15", color=GREEN, numfmt=PRICE)
cell(s, "A5", "WACC (base case)")
cell(s, "C5", "=WACC!C24", color=GREEN, numfmt=PCT)

section(s, "A7", "Intrinsic value per share", "C7")
cell(s, "A8", "Bear case")
cell(s, "C8", f"=DCF!{pershare_cells['BEAR CASE']}", color=GREEN, numfmt=PERS)
cell(s, "A9", "Base case")
cell(s, "C9", f"=DCF!{pershare_cells['BASE CASE']}", color=GREEN, numfmt=PERS, bold=True)
cell(s, "A10", "Bull case")
cell(s, "C10", f"=DCF!{pershare_cells['BULL CASE']}", color=GREEN, numfmt=PERS)
cell(s, "A11", "Comps (average)")
cell(s, "C11", "=Comps!F19", color=GREEN, numfmt=PERS)

section(s, "A13", "Cross-checks", "C13")
cell(s, "A14", "Base-case enterprise value ($mm)")
cell(s, "C14", f"=DCF!C{blocks[1]['hdr']+9}", color=GREEN, numfmt=MONEY)
cell(s, "A15", "Terminal value % of EV (base)")
cell(s, "C15", f"=DCF!C{blocks[1]['hdr']+8}/DCF!C{blocks[1]['hdr']+9}",
     color=GREEN, numfmt=PCT)

# chart data
section(s, "A17", "Valuation range (football field)", "C17")
chart_rows = [("Bear", "C8"), ("Comps", "C11"), ("Base", "C9"), ("Bull", "C10")]
for i, (lbl, ref) in enumerate(chart_rows):
    r = 18 + i
    cell(s, f"A{r}", lbl)
    cell(s, f"C{r}", f"={ref}", color=GREEN, numfmt=PERS)

chart = BarChart()
chart.type = "bar"
chart.title = "Value per share ($) vs reference price"
chart.height = 6
chart.width = 12
data = Reference(s, min_col=3, min_row=18, max_row=21)
cats = Reference(s, min_col=1, min_row=18, max_row=21)
chart.add_data(data, titles_from_data=False)
chart.set_categories(cats)
chart.legend = None
s.add_chart(chart, "E7")

section(s, "A24", "Key caveats", "C24")
caveats = [
    "Standalone Netflix — the pending all-cash Warner Bros. acquisition (Dec-2025)",
    "is NOT modelled and would change debt, shares and the growth profile.",
    "Peer multiples are illustrative placeholders; refresh before citing.",
    "Share price is a post-split reference (10:1, Nov-2025) — update to live quote.",
    "Educational / portfolio use only. Not investment advice.",
]
for i, line in enumerate(caveats):
    cell(s, f"A{25+i}", line, italic=True, size=9)

OUT.parent.mkdir(parents=True, exist_ok=True)
wb.save(OUT)
print(f"Saved {OUT}")
